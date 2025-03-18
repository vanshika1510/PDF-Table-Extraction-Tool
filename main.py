import pdfplumber
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font

# Step 1: Input & Output folders
input_folder = "input"
output_folder = "output"

pdf_files = ['test3.pdf', 'test6.pdf']

for pdf_file in pdf_files:
    print(f"🔍 Processing {pdf_file}...")

    input_path = os.path.join(input_folder, pdf_file)
    output_filename = f"output_for_{os.path.splitext(pdf_file)[0]}.xlsx"
    output_path = os.path.join(output_folder, output_filename)

    combined_rows = []
    table_count = 1
    table_heading_rows = []  # Keep track of heading row numbers

    # Step 2: Extract Tables
    with pdfplumber.open(input_path) as pdf:
        for page in pdf.pages:
            tables_on_page = page.extract_tables({
                "vertical_strategy": "text",
                "horizontal_strategy": "text",
                "intersection_x_tolerance": 10,
                "intersection_y_tolerance": 10
            })
            if tables_on_page:
                for table in tables_on_page:
                    if table:
                        # Add Heading row
                        combined_rows.append([f"Table: {table_count}"])
                        heading_row_num = len(combined_rows)
                        table_heading_rows.append(heading_row_num)
                        
                        # Add table content
                        combined_rows.extend(table)
                        combined_rows.extend([[""] * len(table[0]) for _ in range(10)])  # 10 blank rows
                        table_count += 1

    # Step 3: Save to Excel
    df = pd.DataFrame(combined_rows)
    df.to_excel(output_path, index=False, header=False)

    # Step 4: Format Heading Rows (Bold + Font size 20)
    wb = load_workbook(output_path)
    ws = wb.active
    for row_num in table_heading_rows:
        ws[f"A{row_num}"].font = Font(bold=True, size=20)

    wb.save(output_path)
    print(f"✅ Output saved to {output_path}\n")

print("🎯 All PDFs processed successfully!")
